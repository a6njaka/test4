import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from openpyxl import load_workbook
from matplotlib.ticker import (MultipleLocator, FormatStrFormatter, AutoMinorLocator)

fig1, (ax_a, ax_b) = plt.subplots(figsize=(10, 8), nrows=2, ncols=1)
# fig1.subplots_adjust(left=0.2, bottom=0.2, right=1, top=1, wspace=0.1, hspace=0.1)
folder = r"C:\Users\NJAKA\Desktop\Air pollution RM1/"
excel_file = "Air_Pollution_RM1.xlsx"
wb = load_workbook(excel_file)
ws = wb.active

print(ws["A1"].value)
data = []
x = []
y_pm25 = []
y_aqi_pm25 = []
y_pm10 = []
y_aqi_pm10 = []
US_EPA_STD_PM25 = 50
OMS_STD_PM25 = 25
US_EPA_STD_PM10 = 150
OMS_STD_PM10 = 50
AQI_COLOR = ['#00F600', '#FFFE00', '#FF9900', "#FC3030", "#8A008A", "#990033", "#7D0033"]
aqi_mp25_all = [0, 0, 0, 0, 0, 0, 0]
aqi_mp10_all = [0, 0, 0, 0, 0, 0, 0]
comp_pm25_oms = [0, 0]
comp_pm10_oms = [0, 0]
comp_pm25_epa = [0, 0]
comp_pm10_epa = [0, 0]


def aqi_pm25(c):
    if 0 < c <= 12:
        c_min = 0
        c_max = 12
        i_min = 0
        i_max = 50
        aqi_mp25_all[0] += 1
    elif 12 < c <= 35.4:
        c_min = 12.1
        c_max = 35.4
        i_min = 51
        i_max = 100
        aqi_mp25_all[1] += 1
    elif 35.4 < c <= 55.4:
        c_min = 35.5
        c_max = 55.4
        i_min = 101
        i_max = 150
        aqi_mp25_all[2] += 1
    elif 55.4 < c <= 150.4:
        c_min = 65.5
        c_max = 150.4
        i_min = 151
        i_max = 200
        aqi_mp25_all[3] += 1
    elif 150.4 < c <= 250.4:
        c_min = 150.5
        c_max = 250.4
        i_min = 201
        i_max = 300
        aqi_mp25_all[4] += 1
    elif 250.4 < c <= 350.4:
        c_min = 250.5
        c_max = 500.4
        i_min = 301
        i_max = 400
        aqi_mp25_all[5] += 1
    elif 350.4 < c <= 500.4:
        c_min = 250.5
        c_max = 500.4
        i_min = 401
        i_max = 500
        aqi_mp25_all[6] += 1

    return ((i_max - i_min) / (c_max - c_min)) * (c - c_min) + i_min


def aqi_pm10(c):
    if 0 < c <= 54:
        c_min = 0
        c_max = 54
        i_min = 0
        i_max = 50
        aqi_mp10_all[0] += 1
    elif 54 < c <= 154:
        c_min = 55
        c_max = 154
        i_min = 51
        i_max = 100
        aqi_mp10_all[1] += 1
    elif 154 < c <= 254:
        c_min = 155
        c_max = 254
        i_min = 101
        i_max = 150
        aqi_mp10_all[2] += 1
    elif 254 < c <= 354:
        c_min = 255
        c_max = 354
        i_min = 151
        i_max = 200
        aqi_mp10_all[3] += 1
    elif 354 < c <= 424:
        c_min = 355
        c_max = 424
        i_min = 201
        i_max = 300
        aqi_mp10_all[4] += 1
    elif 424 < c <= 504:
        c_min = 425
        c_max = 504
        i_min = 301
        i_max = 500
        aqi_mp10_all[5] += 1
    elif 504 < c <= 604:
        c_min = 505
        c_max = 604
        i_min = 401
        i_max = 500
        aqi_mp10_all[6] += 1

    return ((i_max - i_min) / (c_max - c_min)) * (c - c_min) + i_min


def zonage_aqi(i):
    if 0 < i < 50:
        return "Bon"
    elif 50 <= i < 100:
        return "Modéré"
    elif 100 <= i < 150:
        return "Malsain pour les personnes sensibles"
    elif 150 <= i < 200:
        return "Malsain"
    elif 200 <= i < 300:
        return "Très malsain"
    elif 300 <= i < 500:
        return "Dangereux"


# ax_a.yaxis.set_major_locator(MultipleLocator(25))
# ax_a.yaxis.set_major_formatter(FormatStrFormatter('%d'))
#
# # For the minor ticks, use no labels; default NullFormatter.
ax_a.xaxis.set_major_locator(MultipleLocator(1))
ax_b.xaxis.set_major_locator(MultipleLocator(1))
label = []

for i in range(3, 51):
    # print([ws[f"B{i}"].value.strftime("%d/%b/%Y"), ws[f"C{i}"].value])
    # x.append(ws[f"B{i}"].value)
    # x.append(ws[f"A{i}"].value.strftime("%d/%b/%Y"))
    x.append(ws[f"A{i}"].value)
    l = ws[f"A{i}"].value
    label.append(f"{l}".replace("2019_2020_RM1_", ""))
    tmp_pm25 = ws[f"D{i}"].value
    tmp_pm10 = ws[f"E{i}"].value + ws[f"D{i}"].value
    y_pm25.append(tmp_pm25)
    y_pm10.append(tmp_pm10)
    tmp_aqi_pm25 = aqi_pm25(tmp_pm25)
    tmp_aqi_pm10 = aqi_pm10(tmp_pm10)
    y_aqi_pm25.append(tmp_aqi_pm25)
    y_aqi_pm10.append(tmp_aqi_pm10)
    ws[f"G{i}"].value = tmp_aqi_pm25
    ws[f"H{i}"].value = tmp_aqi_pm10
    ws[f"I{i}"].value = zonage_aqi(tmp_aqi_pm25)
    ws[f"J{i}"].value = zonage_aqi(tmp_aqi_pm10)

    if tmp_pm25 < OMS_STD_PM25:
        comp_pm25_oms[0] += 1
    else:
        comp_pm25_oms[1] += 1
    if tmp_pm10 < OMS_STD_PM10:
        comp_pm10_oms[0] += 1
    else:
        comp_pm10_oms[1] += 1
    if tmp_pm25 < US_EPA_STD_PM25:
        comp_pm25_epa[0] += 1
    else:
        comp_pm25_epa[1] += 1
    if tmp_pm10 < OMS_STD_PM10:
        comp_pm10_epa[0] += 1
    else:
        comp_pm10_epa[1] += 1
    # data.append([ws[f"A{i}"].value, ws[f"C{i}"].value])
    data.append([ws[f"A{i}"].value, ws[f"C{i}"].value, ws[f"D{i}"].value])

x = []
for i in range(len(y_pm25)):
    x.append(i)
    print(y_pm10[i])

color_pm25 = "#266B07"
color_pm10 = "#3336EC"

ax_a.set_title('Variation du PM2.5 et PM10 à Analakely (Nov 2019 - Jan 2020)')

# plt.subplot(211)
ax_a.set_ylabel('Concentration du PM2.5 (µg/L)', fontsize=10)
ax_a.plot([0, len(data) - 1], [US_EPA_STD_PM25, US_EPA_STD_PM25], color="#FD0128")
ax_a.plot([0, len(data) - 1], [OMS_STD_PM25, OMS_STD_PM25], color="#FD0128")
ax_a.set_ylim(0, 75)
ax_a.text(1, US_EPA_STD_PM25 + 2, f"Normes de l'US EPA pour le PM2.5 ({US_EPA_STD_PM25:0.2f} µg/m³)", ha="left", color="#000000")
ax_a.text(1, OMS_STD_PM25 + 2, f"Normes de l'OMS pour le PM2.5 ({OMS_STD_PM25:0.2f} µg/m³)", ha="left", color="#000000")
ax_a.plot(x, y_pm25, "-o", color=color_pm25, alpha=1, label="PM2.5")
ax_a.legend(loc='upper right')
ax_a.set_xticklabels(["", "", ""] + label, rotation=90, fontsize=8)

# plt.subplot(212)

ax_b.yaxis.set_major_locator(MultipleLocator(50))
ax_b.yaxis.set_major_formatter(FormatStrFormatter('%d'))
ax_b.yaxis.set_minor_locator(MultipleLocator(10))
ax_b.set_ylabel('Concentration du PM10 (µg/L)', fontsize=10)
ax_b.plot([0, len(data) - 1], [US_EPA_STD_PM10, US_EPA_STD_PM10], color="#FD0128", linewidth=1.5)
ax_b.plot([0, len(data) - 1], [OMS_STD_PM10, OMS_STD_PM10], color="#FD0128")
ax_b.set_ylim(0, 200)
ax_b.text(1, US_EPA_STD_PM10 + 2, f"Normes de l'US EPA pour le PM10 ({US_EPA_STD_PM10:0.2f} µg/m³)", ha="left", color="#000000")
ax_b.text(1, OMS_STD_PM10 + 2, f"Normes de l'OMS pour le PM10 ({OMS_STD_PM10:0.2f} µg/m³)", ha="left", color="#000000")
ax_b.plot(x, y_pm10, "-o", color=color_pm10, alpha=1, label="PM10", linewidth=1.5)
ax_b.legend(loc='upper right')
ax_b.set_xticklabels(["", "", ""] + label, rotation=90, fontsize=8)

fig1.savefig(f'images/mp2.5_pm10_std.png')
fig1.savefig(f'images/mp2.5_pm10_std.pdf')

# ------------------------AQI ------------------------
fig2, ax_c = plt.subplots(figsize=(10, 5), nrows=1, ncols=1)
ax_c.set_title('IQA du PM2.5 et PM10 à Analakely (Nov 2019 - Jan 2020)')
ax_c.yaxis.set_minor_locator(MultipleLocator(25))
ax_c.xaxis.set_major_locator(MultipleLocator(1))
ax_c.plot(x, y_aqi_pm25, "-o", color=color_pm25, alpha=1, label="AQI PM2.5", linewidth=1.5)
ax_c.plot(x, y_aqi_pm10, "-o", color=color_pm10, alpha=1, label="AQI PM10", linewidth=1.5)
ax_c.legend(loc='upper right')
ax_c.set_ylim(0, 500)

x = [0, len(x)]
y = [[50, 50], [50, 50], [50, 50], [50, 50], [100, 100], [200, 200]]  # Basic stacked area chart.
ax_c.stackplot(x, y, labels=[], colors=AQI_COLOR)
ax_c.legend(loc='upper right')

ax_c.text(1, 25, "Bon", ha="left", color="#000000", fontsize=10)
ax_c.text(1, 75, "Modéré", ha="left", color="#000000", fontsize=10)
ax_c.text(1, 125, "Malsain pour les personnes sensibles", ha="left", color="#000000", fontsize=10)
ax_c.text(1, 175, "Malsain", ha="left", color="#000000", fontsize=10)
ax_c.text(1, 250, "Très malsain", ha="left", color="#FFFFFF", fontsize=10)
ax_c.text(1, 400, "Dangereux", ha="left", color="#FFFFFF", fontsize=10)
ax_c.set_ylabel('Indice de la Qualité de l’Air (IQA)', fontsize=10)
ax_c.set_xticklabels(["", "", ""] + label, rotation=90, fontsize=8)

fig2.savefig(f'images/aqi_mp2.5_pm10.pdf')
fig2.savefig(f'images/aqi_mp2.5_pm10.png')

print(f"aqi_mp25_all={aqi_mp25_all}")
print(f"aqi_mp10_all={aqi_mp10_all}")

print(f"comp_pm25_oms={comp_pm25_oms}")
print(f"comp_pm10_oms={comp_pm10_oms}")
print(f"comp_pm25_epa={comp_pm25_epa}")
print(f"comp_pm10_epa={comp_pm10_epa}")

wb.save((excel_file))

plt.show()
