import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from openpyxl import load_workbook
from matplotlib.ticker import (MultipleLocator, FormatStrFormatter, AutoMinorLocator)

fig, ax = plt.subplots(figsize=(15, 6))

excel_file = "RA7016-PM2.5&PM10-Madagascar.xlsx"
wb = load_workbook(excel_file)
ws = wb.active

print(ws["A1"].value)
data = []
x = []
y_pm25 = []
y_pm10 = []

ax.yaxis.set_major_locator(MultipleLocator(25))
ax.yaxis.set_major_formatter(FormatStrFormatter('%d'))

# For the minor ticks, use no labels; default NullFormatter.
ax.yaxis.set_minor_locator(MultipleLocator(5))

for i in range(2, 106):
    print([ws[f"A{i}"].value.strftime("%d/%b/%Y"), ws[f"C{i}"].value])
    # x.append(ws[f"B{i}"].value)
    x.append(ws[f"A{i}"].value.strftime("%d/%b/%Y"))
    y_pm25.append(ws[f"C{i}"].value)
    y_pm10.append(ws[f"D{i}"].value)
    # data.append([ws[f"A{i}"].value, ws[f"C{i}"].value])
    data.append([ws[f"A{i}"].value.strftime("%d/%b/%Y"), ws[f"C{i}"].value, ws[f"D{i}"].value])

US_EAP_STD_PM25 = 50  # µg/m3
US_EAP_STD_PM10 = 150  # µg/m3
plt.plot([0, len(data) - 1], [US_EAP_STD_PM25, US_EAP_STD_PM25], color="#FD0128")
plt.plot([0, len(data) - 1], [US_EAP_STD_PM10, US_EAP_STD_PM10], color="#A90120")
# plt.text(1, US_EAP_STD_PM25 + 2, f"US_EAP_STD_PM25 ({US_EAP_STD_PM25:0.2f} µg/m3)", ha="left", color="#FD0128")
# plt.text(1, US_EAP_STD_PM10 + 2, f"US_EAP_STD_PM25 ({US_EAP_STD_PM10:0.2f} µg/m3)", ha="left", color="#FD0128")
plt.text(1, US_EAP_STD_PM25 + 2, f"US_EAP_STD_PM25 ({US_EAP_STD_PM25:0.2f} µg/m³)", ha="left", color="#000000")
plt.text(1, US_EAP_STD_PM10 + 2, f"US_EAP_STD_PM25 ({US_EAP_STD_PM10:0.2f} µg/m³)", ha="left", color="#000000")


# plt.ylim(0, 1)
plt.plot(x, y_pm25, '-ob', dash_capstyle="round", alpha=1, label="PM2.5")
# plt.plot(x, y_pm25, 'ob', alpha=0.5)
plt.plot(x, y_pm10, '-o', dash_capstyle="round", alpha=1, color="#9200AC", label="PM10")
# plt.plot(x, y_pm10, 'o', alpha=0.5, color="#9200AC")
plt.title("AQI PM2.5 - INSTN-Madagascar 2017-2018")
# plt.xlabel("Standard (µg/L)")
plt.ylabel("Air Quality Index (AQI)")
# plt.legend(loc='lower right')
# plt.grid(color='#EBCAFE')
# ax.yaxis.grid(color='#EBCAFE')
plt.get_default_filename = lambda: 'new_default_name.png'
plt.xlim(0, len(data))
# plt.ylim(0, 30)

# i = 0
# for tt in data:
#     plt.text(i, tt[1], f"{tt[1]:0.2f}", ha="center")
#     plt.text(i, tt[2], f"{tt[2]:0.2f}", ha="center")
#     i += 1

AQI_Regions = (
    (150.4, 250, "Very Unhealthy", "#99004C"),
    (55.4, 150.4, "Unhealthy", "#FF0000"),
    (35.4, 55.4, "Unhealthy for Sensitive Groups", "#FF7E00"),
    (12, 35.4, "Moderate", "#FFFF00"),
    (0, 12, "Good", "#00E400"),
)

# for aqi in AQI_Regions:
#     plt.fill_between([0, len(data) - 1], [aqi[0], aqi[0]], [aqi[1], aqi[1]], alpha=01.0, label=aqi[2], color=aqi[3])
# plt.text(0, aqi[0], aqi[2])
plt.xticks(rotation=90)
# plt.legend(loc='upper left')
# plt.margins(0.71)
# plt.subplots_adjust(bottom=0.15)
plt.subplots_adjust(bottom=0.25)
# ax.annotate('Outlier', xy=(5, 0.4), xytext=(3, 0.7), arrowprops=dict(facecolor='black', shrink=0.0000025), )
# ax.annotate('Outlier', xy=(5, 0.4), xytext=(8, 0.4), arrowprops=dict(arrowstyle="->", connectionstyle="arc3"), )
plt.legend(loc='upper right')
plt.ylim(0, 200)

output_folder = r"C:\Users\NJAKA\Desktop"
# plt.savefig(f'{output_folder}/matplotlib.png')
# plt.savefig(f'{output_folder}/matplotlib.pdf')
plt.show()
