import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from openpyxl import load_workbook
from matplotlib.ticker import (MultipleLocator, FormatStrFormatter, AutoMinorLocator)

fig, (ax_a, ax_b) = plt.subplots(figsize=(10, 8), nrows=2, ncols=1)

folder = r"C:\Users\NJAKA\Desktop\Air pollution RM1/"
excel_file = "Air_Pollution_RM1.xlsx"
wb = load_workbook(folder + excel_file)
ws = wb.active

print(ws["A1"].value)
data = []
x = []
y_pm25 = []
y_pm10 = []
US_EPA_STD_PM25 = 50
OMS_STD_PM25 = 25
US_EPA_STD_PM10 = 150
OMS_STD_PM10 = 50


def aqi_pm25(c):
    if 0 < c <= 15.4:
        c_min = 0
        c_max = 14.5
        i_min = 0
        i_max = 50
    elif 15.4 < c <= 40.4:
        c_min = 15.5
        c_max = 40.4
        i_min = 51
        i_max = 100
    elif 40.4 < c <= 65.4:
        c_min = 40.5
        c_max = 65.4
        i_min = 101
        i_max = 150
    elif 65.4 < c <= 150.4:
        c_min = 65.5
        c_max = 150.4
        i_min = 151
        i_max = 200
    elif 150.4 < c <= 250.4:
        c_min = 150.5
        c_max = 250.4
        i_min = 201
        i_max = 300
    elif 250.4 < c <= 500.4:
        c_min = 250.5
        c_max = 500.4
        i_min = 301
        i_max = 500

    return ((i_max - i_min) / (c_max - c_min)) * (c - c_min) + i_min


def aqi_pm10(c):
    if 0 < c <= 54:
        c_min = 0
        c_max = 54
        i_min = 0
        i_max = 50
    elif 54 < c <= 154:
        c_min = 55
        c_max = 154
        i_min = 51
        i_max = 100
    elif 154 < c <= 254:
        c_min = 155
        c_max = 254
        i_min = 101
        i_max = 150
    elif 254 < c <= 354:
        c_min = 255
        c_max = 354
        i_min = 151
        i_max = 200
    elif 354 < c <= 424:
        c_min = 355
        c_max = 424
        i_min = 201
        i_max = 300
    elif 424 < c <= 604:
        c_min = 425
        c_max = 604
        i_min = 301
        i_max = 500

    return ((i_max - i_min) / (c_max - c_min)) * (c - c_min) + i_min


ax_a.yaxis.set_major_locator(MultipleLocator(25))
ax_a.yaxis.set_major_formatter(FormatStrFormatter('%d'))

# For the minor ticks, use no labels; default NullFormatter.
ax_a.yaxis.set_minor_locator(MultipleLocator(5))
label = []

for i in range(3, 51):
    # print([ws[f"B{i}"].value.strftime("%d/%b/%Y"), ws[f"C{i}"].value])
    # x.append(ws[f"B{i}"].value)
    # x.append(ws[f"A{i}"].value.strftime("%d/%b/%Y"))
    x.append(ws[f"A{i}"].value)
    l = ws[f"A{i}"].value
    label.append(f"{l}".replace("2019_2020_RM1_", ""))
    y_pm25.append(ws[f"D{i}"].value)
    y_pm10.append(ws[f"E{i}"].value + ws[f"D{i}"].value)
    # data.append([ws[f"A{i}"].value, ws[f"C{i}"].value])
    data.append([ws[f"A{i}"].value, ws[f"C{i}"].value, ws[f"D{i}"].value])

x = []
for i in range(len(y_pm25)):
    x.append(i)
    print(y_pm10[i])

plt.subplot(211)
plt.ylabel('Concentration du PM2.5 (µg/L)', fontsize=10)
plt.plot(x, y_pm25, "-o", color="green", alpha=1, label="PM2.5")
plt.plot([0, len(data) - 1], [US_EPA_STD_PM25, US_EPA_STD_PM25], color="#FD0128")
plt.plot([0, len(data) - 1], [OMS_STD_PM25, OMS_STD_PM25], color="#FD0128")
plt.ylim(-10, 75)
plt.text(1, US_EPA_STD_PM25 + 2, f"US_EPA_STD_PM25 ({US_EPA_STD_PM25:0.2f} µg/m³)", ha="left", color="#000000")
plt.text(1, OMS_STD_PM25 + 2, f"OMS_STD_PM25 ({OMS_STD_PM25:0.2f} µg/m³)", ha="left", color="#000000")
plt.legend(loc='upper right')
plt.suptitle('PM2.5 et PM10 à Analakely (Nov 2019 - Jan 2020)')
plt.xticks(x, label)
plt.xticks(rotation=90, fontsize=8)
# plt.xlabel('sample code', fontsize=10)

plt.subplot(212)
plt.plot(x, y_pm10, "-o", color="#9E0AEC", alpha=1, label="PM10", linewidth=1.5)
plt.ylabel('Concentration du PM10 (µg/L)', fontsize=10)
plt.plot([0, len(data) - 1], [US_EPA_STD_PM10, US_EPA_STD_PM10], color="#FD0128", linewidth=1.5)
plt.plot([0, len(data) - 1], [OMS_STD_PM10, OMS_STD_PM10], color="#FD0128")
plt.ylim(-20, 200)
plt.text(1, US_EPA_STD_PM10 + 2, f"US_EPA_STD_PM10 ({US_EPA_STD_PM10:0.2f} µg/m³)", ha="left", color="#000000")
plt.text(1, OMS_STD_PM10 + 2, f"OMS_STD_PM10 ({OMS_STD_PM10:0.2f} µg/m³)", ha="left", color="#000000")
plt.legend(loc='upper right')
ax_a.set_title('Contour Plot')
plt.xticks(x, label)
plt.xticks(rotation=90, fontsize=8)

plt.savefig(f'{folder}images/mp2.5_pm10_std.png')
plt.savefig(f'{folder}images/mp2.5_pm10_std.pdf')

plt.show()
